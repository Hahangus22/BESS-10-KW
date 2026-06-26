import openpyxl
import os

file_path = r"c:\Users\LENOVO\Documents\PlatformIO\Projects\Layar BESS 10Kw jual\src\BMS Modbus RTU 2.xlsx"
if os.path.exists(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    for sheet in wb.sheetnames:
        print(f"Sheet: {sheet}")
        ws = wb[sheet]
        for r in range(1, 100):
            row_vals = [ws.cell(r, c).value for c in range(1, 15)]
            if any(row_vals):
                print(f"Row {r}: {row_vals}")
else:
    print("File not found")
